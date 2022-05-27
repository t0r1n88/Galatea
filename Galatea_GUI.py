import tkinter
import numpy as np
import sys
import pandas as pd
from docxtpl import DocxTemplate
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
pd.options.mode.chained_assignment = None  # default='warn'
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from itertools import islice
import time
import os

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def change_name_discipline(cell):
    # Функция для установки заглавной буквы для содержимого ячейки
    # очищаем от пробельных символов спереди и сзади
    value = cell.strip()
    # Делаем заглавным первый символ
    out_value = f'{value[0].upper()}{value[1:]}'

    return out_value


def select_folder_data():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data
    path_folder_data = filedialog.askdirectory()

def select_end_folder():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()

def select_file_docx():
    """
    Функция для выбора файла Word
    :return: Путь к файлу шаблона
    """
    global file_docx
    file_docx = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))

def select_file_data_xlsx():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global file_data_xlsx
    # Получаем путь к файлу
    file_data_xlsx = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_files_data_xlsx():
    """
    Функция для выбора нескоьких файлов с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global files_data_xlsx
    # Получаем путь файлы
    files_data_xlsx = filedialog.askopenfilenames(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))



def processing_data():
    """
    Фугкция для обработки данных
    :return:
    """
    # Создаем общий датафрейм
    try:
        col_out_df = ['Ф.И.О. преподавателя (полностью)', 'Занимаемая в ПОО должность', 'Квалификационная категория','Ученная степень, звание'
                      ,'Учебная дисциплина',
                      'Курс', 'Группа', 'Теория', 'ЛПЗ', 'Учебная практика', 'Производственная практика',
                      'Преддипломная практика', 'Руководство над курсовым проектом'
            , 'Консультации', 'Контроль (экзамены, зачеты и т.д.)', 'Руководство ВКР', 'ИТОГО', 'Итого по тарификации']
        out_df = pd.DataFrame(columns=col_out_df)

        # Обрабатываем файл, пропускаем все не xlsx файлы и временные файлы
        for file in files_data_xlsx:
            if not file.startswith('~') and file.endswith('.xlsx'):
                # Создаем датафрейм для соединения результата обработки таблицы и строки с суммой
                finish_df = pd.DataFrame()
                df = pd.read_excel(file, skiprows=6)
                df.columns = ['№/пп', 'Наименование группы', 'Наименование дисциплины', '1 семестр кол 1',
                              '1 семестр кол 2',
                              '2 семестр кол 1', '2 семестр кол 2', 'Всего часов',
                              'Теория', 'ЛПЗ', 'Учебная практика', 'Производственная практика', 'Преддипломная практика',
                              'Руководство над курсовым проектом', 'Консультации', 'Контроль (экзамены, зачеты и т.д.)',
                              'Руководство ВКР', 'ИТОГО', 'Преподаватель', 'Промежуточные суммы']

                df = df[df['Наименование группы'].notna()]

                # Очищаем от пробелов перед и после слов
                df['Наименование группы'] = df['Наименование группы'].apply(lambda x: x.strip())
                df['Наименование дисциплины'] = df['Наименование дисциплины'].apply(
                    lambda x: change_name_discipline(x) if type(x) == str else x)

                df = df[df['Наименование группы'] != 'ознакомлен']

                # Удаляем лишний столбец
                df.drop(columns=['Промежуточные суммы'], inplace=True)
                # Создаем книгу для того чтобы отобрать все строки до внебюджета
                wb = openpyxl.Workbook()
                ws = wb.active

                for r in dataframe_to_rows(df, index=True, header=True):
                    if 'внебюджет' in r:
                        break
                    ws.append(r)

                # Загружаем обратно очищенный от внебюджетных дисциплин датафрейм
                data = ws.values
                cols = next(data)[1:]
                data = list(data)
                idx = [r[0] for r in data]
                data = (islice(r, 1, None) for r in data)
                clear_df = pd.DataFrame(data, index=idx, columns=cols)

                clear_df.sort_values(by='Наименование дисциплины', inplace=True)

                # Копируем данные  в датафрейм
                finish_df['Ф.И.О. преподавателя (полностью)'] = clear_df['Преподаватель']
                finish_df['Занимаемая в ПОО должность'] = ''
                finish_df['Квалификационная категория'] = ''
                finish_df['Ученная степень, звание'] = ''
                finish_df['Учебная дисциплина'] = clear_df['Наименование дисциплины']
                finish_df['Курс'] = ''
                finish_df['Группа'] = clear_df['Наименование группы']
                finish_df['Теория'] = clear_df['Теория']
                finish_df['ЛПЗ'] = clear_df['ЛПЗ']
                finish_df['Учебная практика'] = clear_df['Учебная практика']
                finish_df['Производственная практика'] = clear_df['Производственная практика']
                finish_df['Преддипломная практика'] = clear_df['Преддипломная практика']
                finish_df['Руководство над курсовым проектом'] = clear_df['Руководство над курсовым проектом']
                finish_df['Консультации'] = clear_df['Консультации']
                finish_df['Контроль (экзамены, зачеты и т.д.)'] = clear_df['Контроль (экзамены, зачеты и т.д.)']
                finish_df['Руководство ВКР'] = clear_df['Руководство ВКР']
                finish_df['ИТОГО'] = clear_df['ИТОГО']
                finish_df['Итого по тарификации'] = ''

                finish_df.dropna(subset=['Ф.И.О. преподавателя (полностью)'], inplace=True)

                # Получаем сумму колонок
                sum_col = finish_df.sum(axis=0, numeric_only=True).to_frame().T
                sum_col['Ф.И.О. преподавателя (полностью)'] = 'Итого'

                finish_df = pd.concat([finish_df, sum_col], ignore_index=True)

                out_df = pd.concat([out_df, finish_df], ignore_index=True)

        # Создаем книгу для итогового файла
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active

        # Записываем финальный датафрейм в созданную книгу
        for r in dataframe_to_rows(out_df, index=False, header=True):
            if len(r) != 1:
                out_ws.append(r)
        for cell in out_ws[1]:
            cell.style = 'Headline 4'
            cell.alignment = Alignment(wrap_text=True)
        # выделяем строки с Итого
        number_row_itog_lst = []
        # Создаем счетчик
        count_row = 0
        for cell in out_ws['A']:
            count_row += 1
            if cell.value == 'Итого':
                number_row_itog_lst.append(count_row)

        for number_row in number_row_itog_lst:
            for cell in out_ws[number_row]:
                cell.style = 'Total'

        out_wb['Sheet'].column_dimensions['A'].width = 35
        out_wb['Sheet'].column_dimensions['B'].width = 15
        out_wb['Sheet'].column_dimensions['C'].width = 15
        out_wb['Sheet'].column_dimensions['E'].width = 50




        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        out_wb.save(f'{path_to_end_folder}/Приложение №6 от {current_time}.xlsx')
    except NameError:
        messagebox.showerror('ЦОПП Бурятия', f'Выберите файл с данными и папку куда будет генерироваться файл')
    except ValueError:
        messagebox.showerror('ЦОПП Бурятия', f'Проверьте количество колонок в файле {file}')
    except KeyError:
        messagebox.showerror('ЦОПП Бурятия', f'Проверьте названия колонок в файле{file}')

    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Данные успешно обработаны')


if __name__ == '__main__':
    window = Tk()
    window.title('ЦОПП Бурятия')
    window.geometry('700x860')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку обработки данных для Приложения 6
    tab_report_6 = ttk.Frame(tab_control)
    tab_control.add(tab_report_6, text='Приложение №6')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_report_6,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nОбработка данных для приложения №6')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img = PhotoImage(file=path_to_img)
    Label(tab_report_6,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data = Button(tab_report_6, text='1) Выберите файлы с данными', font=('Arial Bold', 20),
                          command=select_files_data_xlsx
                          )
    btn_choose_data.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_report_6, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder
                                       )
    btn_choose_end_folder.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку обработки данных

    btn_proccessing_data = Button(tab_report_6, text='3) Обработать данные', font=('Arial Bold', 20),
                                       command=processing_data
                                       )
    btn_proccessing_data.grid(column=0, row=4, padx=10, pady=10)

    window.mainloop()